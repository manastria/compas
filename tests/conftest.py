"""Fixtures partagées pour les tests Compas."""

from pathlib import Path

import openpyxl
import pytest


def make_workbook() -> openpyxl.Workbook:
    """Crée un classeur xlsx de test représentatif.

    Contient :
    - Feuille Config : 3 étudiants (Alice avec INE, Bob anonyme, Eve avec date de départ)
    - Feuille S1 : séance 1 avec les 3 étudiants (présences variées, scores, heure_fin)
    - Feuille S2 : séance 2 avec Alice uniquement
    - Feuille Modèle : ignorée
    - Feuille tmp-brouillon : ignorée
    """
    wb = openpyxl.Workbook()

    # --- Feuille Config ---
    # Colonnes : A=Nom, B=INE, C=Anonyme, D=Pseudo, E=Date de départ
    config = wb.active
    config.title = "Config"
    config["A1"] = "Projet"
    config["B1"] = "Infrastructure réseau PME"
    config["A2"] = "Groupe"
    config["B2"] = "TP1"
    # En-têtes ligne 4
    config.cell(row=4, column=1, value="Nom")
    config.cell(row=4, column=2, value="INE")
    config.cell(row=4, column=3, value="Anonyme")
    config.cell(row=4, column=4, value="Pseudo")
    config.cell(row=4, column=5, value="Date de départ")
    # Étudiants (ligne 5+)
    config.cell(row=5, column=1, value="Dupont Alice")
    config.cell(row=5, column=2, value="INE001")
    config.cell(row=5, column=3, value="non")
    config.cell(row=6, column=1, value="Martin Bob")
    config.cell(row=6, column=2, value="INE002")
    config.cell(row=6, column=3, value="oui")
    config.cell(row=6, column=4, value="BobM")
    config.cell(row=7, column=1, value="Leclerc Eve")
    config.cell(row=7, column=3, value="non")
    config.cell(row=7, column=5, value="2026-02-15")

    # --- Feuille S1 ---
    s1 = wb.create_sheet("S1")
    s1["A1"] = "COMPAS — Relevé de séance"
    s1["A2"] = "Séance n°"
    s1["B2"] = 1
    s1["C2"] = "Date"
    s1["D2"] = "15/01/2026"
    s1["E2"] = "Heure début"
    s1["F2"] = "8h00"
    s1["G2"] = "Enseignant"
    s1["H2"] = "Prof Martin"
    s1["I2"] = "Heure fin"
    s1["J2"] = "12h00"
    s1.cell(row=4, column=1, value="Étudiant")
    s1.cell(row=5, column=1, value="(symboles)")
    # Données étudiants
    s1.cell(row=6, column=1, value="Dupont Alice")
    s1.cell(row=6, column=2, value="P")
    s1.cell(row=6, column=3, value=1)    # autonomie
    s1.cell(row=6, column=4, value=2)    # rigueur
    s1.cell(row=6, column=5, value=0)    # communication
    s1.cell(row=6, column=6, value=1)    # engagement
    s1.cell(row=6, column=7, value="Bon travail")
    s1.cell(row=7, column=1, value="Martin Bob")
    s1.cell(row=7, column=2, value="R:15")
    s1.cell(row=7, column=3, value=0)    # autonomie
    s1.cell(row=7, column=4, value=None) # rigueur : non observé
    s1.cell(row=7, column=5, value=1)    # communication
    s1.cell(row=7, column=6, value=-1)   # engagement
    s1.cell(row=8, column=1, value="Leclerc Eve")
    s1.cell(row=8, column=2, value="A")  # absent

    # --- Feuille S2 ---
    s2 = wb.create_sheet("S2")
    s2["A2"] = "Séance n°"
    s2["B2"] = 2
    s2["D2"] = "22/01/2026"
    s2["F2"] = "8h00"
    s2["H2"] = "Prof Martin"
    s2["J2"] = "12h00"
    s2.cell(row=6, column=1, value="Dupont Alice")
    # presence vide → sera interprété comme P
    s2.cell(row=6, column=3, value=2)    # autonomie
    s2.cell(row=6, column=4, value=1)    # rigueur
    s2.cell(row=6, column=5, value=1)    # communication
    s2.cell(row=6, column=6, value=2)    # engagement

    # --- Feuilles ignorées ---
    wb.create_sheet("Modèle")
    wb.create_sheet("tmp-brouillon")

    return wb


@pytest.fixture
def test_xlsx(tmp_path: Path) -> Path:
    """Fichier xlsx de test dans un répertoire temporaire."""
    data_dir = tmp_path / "data"
    data_dir.mkdir()
    xlsx_path = data_dir / "test_projet.xlsx"
    make_workbook().save(xlsx_path)
    return xlsx_path


@pytest.fixture
def test_db_path(tmp_path: Path) -> Path:
    """Chemin vers une base SQLite temporaire (répertoire créé, fichier absent)."""
    return tmp_path / "output" / "compas.db"
