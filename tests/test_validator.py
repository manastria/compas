"""Tests du module compas.validator."""

import openpyxl
import pytest

from compas.validator import Issue, Severity, _invalid_presence_tokens, validate_xlsx
from tests.conftest import make_workbook


def _issues(wb: openpyxl.Workbook, tmp_path) -> list[Issue]:
    """Sauvegarde un workbook temporaire et lance la validation."""
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return validate_xlsx(path)


def _errors(issues: list[Issue]) -> list[Issue]:
    return [i for i in issues if i.severity == Severity.ERROR]


def _warnings(issues: list[Issue]) -> list[Issue]:
    return [i for i in issues if i.severity == Severity.WARNING]


# ---------------------------------------------------------------------------
# Fichier valide
# ---------------------------------------------------------------------------


class TestFichierValide:
    def test_workbook_conforme_aucun_probleme(self, tmp_path):
        issues = _issues(make_workbook(), tmp_path)
        assert issues == []

    def test_fichier_genere_aucun_probleme(self, tmp_path):
        """Le fichier généré par generate_test_data doit être conforme."""
        path = tmp_path.parent / "test_generated.xlsx"
        if not path.exists():
            pytest.skip("Fichier test_generated.xlsx absent — lancez generate_test_data.py d'abord")
        issues = validate_xlsx(path)
        assert _errors(issues) == []


# ---------------------------------------------------------------------------
# Erreurs structurelles
# ---------------------------------------------------------------------------


class TestErreursFichier:
    def test_fichier_inexistant(self, tmp_path):
        issues = validate_xlsx(tmp_path / "inexistant.xlsx")
        assert len(_errors(issues)) == 1
        assert "Impossible d'ouvrir" in issues[0].message

    def test_config_absente(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "S1"
        issues = _issues(wb, tmp_path)
        assert any("Config" in i.message for i in _errors(issues))


# ---------------------------------------------------------------------------
# Validation Config
# ---------------------------------------------------------------------------


class TestValidationConfig:
    def test_b1_vide_erreur(self, tmp_path):
        wb = make_workbook()
        wb["Config"]["B1"] = None
        issues = _issues(wb, tmp_path)
        assert any("B1" in i.location for i in _errors(issues))

    def test_b2_vide_avertissement(self, tmp_path):
        wb = make_workbook()
        wb["Config"]["B2"] = None
        issues = _issues(wb, tmp_path)
        assert any("B2" in i.location for i in _warnings(issues))

    def test_ine_format_non_standard_avertissement(self, tmp_path):
        wb = make_workbook()
        wb["Config"].cell(row=5, column=2).value = "INE_INVALIDE"  # format non standard
        issues = _issues(wb, tmp_path)
        w = [i for i in _warnings(issues) if "INE" in i.message]
        assert len(w) >= 1
        assert any("INE_INVALIDE" in i.message for i in w)

    def test_ine_format_standard_pas_davertissement(self, tmp_path):
        # Le conftest utilise déjà des INE standard — aucun avertissement INE attendu
        issues = _issues(make_workbook(), tmp_path)
        assert not any("INE" in i.message for i in issues)

    def test_anonyme_invalide_erreur(self, tmp_path):
        wb = make_workbook()
        wb["Config"].cell(row=5, column=3, value="maybe")
        issues = _issues(wb, tmp_path)
        assert any("Anonyme" in i.message for i in _errors(issues))

    def test_anonyme_oui_sans_pseudo_avertissement(self, tmp_path):
        wb = make_workbook()
        wb["Config"].cell(row=5, column=3, value="oui")
        wb["Config"].cell(row=5, column=4, value=None)
        issues = _issues(wb, tmp_path)
        assert any("anonyme sans pseudo" in i.message for i in _warnings(issues))

    def test_anonyme_oui_avec_pseudo_ok(self, tmp_path):
        wb = make_workbook()
        wb["Config"].cell(row=5, column=3, value="oui")
        wb["Config"].cell(row=5, column=4, value="alice42")
        issues = _issues(wb, tmp_path)
        assert not any("anonyme sans pseudo" in i.message for i in issues)

    def test_date_depart_invalide_erreur(self, tmp_path):
        wb = make_workbook()
        wb["Config"].cell(row=7, column=5, value="pas-une-date")
        issues = _issues(wb, tmp_path)
        assert any("Date de départ" in i.message for i in _errors(issues))

    def test_aucun_etudiant_avertissement(self, tmp_path):
        # Construire un classeur Config sans étudiants (pas de make_workbook pour éviter
        # les cellules déjà remplies — openpyxl ne clear pas cell(value=None) en place)
        wb = openpyxl.Workbook()
        config = wb.active
        config.title = "Config"
        config["A1"] = "Projet"
        config["B1"] = "Test"
        config["A2"] = "Groupe"
        config["B2"] = "G1"
        config.cell(row=4, column=1, value="Nom")
        # Ligne 5 intentionnellement vide → aucun étudiant
        issues = _issues(wb, tmp_path)
        assert any("Aucun étudiant" in i.message for i in _warnings(issues))


# ---------------------------------------------------------------------------
# Validation feuilles de séance
# ---------------------------------------------------------------------------


class TestValidationSeance:
    def test_b2_manquant_erreur(self, tmp_path):
        wb = make_workbook()
        wb["S1"]["B2"] = None
        issues = _issues(wb, tmp_path)
        assert any("S1" in i.sheet and "B2" in i.location for i in _errors(issues))

    def test_b2_non_entier_erreur(self, tmp_path):
        wb = make_workbook()
        wb["S1"]["B2"] = "un"
        issues = _issues(wb, tmp_path)
        assert any("non entier" in i.message and "S1" in i.sheet for i in _errors(issues))

    def test_date_manquante_erreur(self, tmp_path):
        wb = make_workbook()
        wb["S1"]["D2"] = None
        issues = _issues(wb, tmp_path)
        assert any("D2" in i.location and "S1" in i.sheet for i in _errors(issues))

    def test_date_invalide_erreur(self, tmp_path):
        wb = make_workbook()
        wb["S1"]["D2"] = "32/13/2026"
        issues = _issues(wb, tmp_path)
        assert any("D2" in i.location and "S1" in i.sheet for i in _errors(issues))

    def test_heure_debut_manquante_avertissement(self, tmp_path):
        wb = make_workbook()
        wb["S1"]["F2"] = None
        issues = _issues(wb, tmp_path)
        assert any("F2" in i.location and "S1" in i.sheet for i in _warnings(issues))

    def test_enseignant_manquant_avertissement(self, tmp_path):
        wb = make_workbook()
        wb["S1"]["H2"] = None
        issues = _issues(wb, tmp_path)
        assert any("H2" in i.location and "S1" in i.sheet for i in _warnings(issues))

    def test_heure_fin_manquante_avertissement(self, tmp_path):
        wb = make_workbook()
        wb["S1"]["J2"] = None
        issues = _issues(wb, tmp_path)
        assert any("J2" in i.location and "S1" in i.sheet for i in _warnings(issues))

    def test_etudiant_inconnu_avertissement(self, tmp_path):
        wb = make_workbook()
        wb["S1"].cell(row=9, column=1, value="Inconnu Tartempion")
        issues = _issues(wb, tmp_path)
        assert any("Tartempion" in i.message for i in _warnings(issues))

    def test_score_hors_plage_erreur(self, tmp_path):
        wb = make_workbook()
        wb["S1"].cell(row=6, column=3, value=5)  # autonomie = 5
        issues = _issues(wb, tmp_path)
        assert any("hors plage" in i.message and "S1" in i.sheet for i in _errors(issues))

    def test_score_non_entier_erreur(self, tmp_path):
        wb = make_workbook()
        wb["S1"].cell(row=6, column=4, value="bon")  # rigueur = "bon"
        issues = _issues(wb, tmp_path)
        assert any("non entier" in i.message and "S1" in i.sheet for i in _errors(issues))

    def test_score_null_ok(self, tmp_path):
        wb = make_workbook()
        wb["S1"].cell(row=6, column=3, value=None)
        issues = _issues(wb, tmp_path)
        assert not any("hors plage" in i.message or "non entier" in i.message for i in issues)

    def test_doublons_seance_avertissement(self, tmp_path):
        wb = make_workbook()
        # S2 utilise le même numéro que S1
        wb["S2"]["B2"] = 1
        issues = _issues(wb, tmp_path)
        assert any("déjà utilisé" in i.message for i in _warnings(issues))

    def test_aucune_seance_avertissement(self, tmp_path):
        wb = openpyxl.Workbook()
        config = wb.active
        config.title = "Config"
        config["A1"] = "Projet"
        config["B1"] = "Test"
        config["A2"] = "Groupe"
        config["B2"] = "G1"
        config.cell(row=4, column=1, value="Nom")
        config.cell(row=5, column=1, value="Alice")
        config.cell(row=5, column=3, value="non")
        issues = _issues(wb, tmp_path)
        assert any("Aucune feuille de séance" in i.message for i in _warnings(issues))


# ---------------------------------------------------------------------------
# Validation présence
# ---------------------------------------------------------------------------


class TestValidationPresence:
    @pytest.mark.parametrize("value", [
        "P", "A", "N",
        "A:medical", "A:9h15-10h00", "A:9h15-10h00:medical",
        "A:H1-H2", "A:H1-H2:sport",
        "R:15", "R:9h30", "R:9h30:transport",
        "RR:10", "RR:10:famille",
        "D:10h15", "D:10h15:medical",
        "R:5,RR:10", "R:10,D:11h30",
    ])
    def test_presence_valide(self, value):
        assert _invalid_presence_tokens(value) == []

    @pytest.mark.parametrize("value", [
        "X", "Z:15", "retard", "absent",
    ])
    def test_presence_invalide(self, value):
        assert _invalid_presence_tokens(value) != []

    def test_presence_vide_ok(self):
        assert _invalid_presence_tokens("") == []

    def test_presence_dans_seance(self, tmp_path):
        wb = make_workbook()
        wb["S1"].cell(row=6, column=2, value="INCONNU:truc")
        issues = _issues(wb, tmp_path)
        assert any("Token de présence" in i.message and "S1" in i.sheet for i in _warnings(issues))
