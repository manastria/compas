"""Tests d'intégration de bout en bout : import multi-fichiers → dashboard HTML."""

import json
import re
import sqlite3

import openpyxl
import pytest

from compas.dashboard import generate
from compas.importer import import_all
from tests.conftest import make_workbook


def _make_second_workbook() -> openpyxl.Workbook:
    """Crée un second classeur avec Dupont Alice (partagée) et Lemaire Paul (unique)."""
    wb = openpyxl.Workbook()
    config = wb.active
    config.title = "Config"
    config["A1"] = "Projet"
    config["B1"] = "Développement Web"
    config["A2"] = "Groupe"
    config["B2"] = "TP2"
    config.cell(row=4, column=1, value="Nom")
    config.cell(row=5, column=1, value="Dupont Alice")
    config.cell(row=5, column=2, value="INE001")
    config.cell(row=5, column=3, value="non")
    config.cell(row=6, column=1, value="Lemaire Paul")
    config.cell(row=6, column=3, value="non")

    s1 = wb.create_sheet("S1")
    s1["A2"] = "Séance n°"
    s1["B2"] = 1
    s1["D2"] = "01/02/2026"
    s1["F2"] = "8h00"
    s1["H2"] = "Prof Dupont"
    s1["J2"] = "12h00"
    s1.cell(row=6, column=1, value="Dupont Alice")
    s1.cell(row=6, column=2, value="P")
    s1.cell(row=6, column=3, value=2)   # autonomie
    s1.cell(row=6, column=4, value=1)   # rigueur
    s1.cell(row=6, column=5, value=1)   # communication
    s1.cell(row=6, column=6, value=2)   # engagement
    s1.cell(row=7, column=1, value="Lemaire Paul")
    s1.cell(row=7, column=2, value="R:10")
    s1.cell(row=7, column=3, value=1)   # autonomie

    return wb


def _extract_compas_json(html: str) -> dict:
    m = re.search(r"var COMPAS_DATA = ({.*?});", html, re.DOTALL)
    assert m, "var COMPAS_DATA introuvable dans le HTML généré"
    return json.loads(m.group(1))


# ---------------------------------------------------------------------------
# Tests end-to-end : un seul fichier
# ---------------------------------------------------------------------------


class TestPipelineFichierUnique:
    def test_html_existe(self, test_xlsx, test_db_path, tmp_path):
        import_all(test_xlsx.parent, test_db_path)
        out = tmp_path / "dashboard.html"
        generate(test_db_path, out)
        assert out.exists()

    def test_json_projet_correct(self, test_xlsx, test_db_path, tmp_path):
        import_all(test_xlsx.parent, test_db_path)
        out = tmp_path / "dashboard.html"
        generate(test_db_path, out)
        data = _extract_compas_json(out.read_text(encoding="utf-8"))
        assert data["projet"] == "Infrastructure réseau PME"
        assert data["groupe"] == "TP1"

    def test_etudiants_actifs_presents(self, test_xlsx, test_db_path, tmp_path):
        """Alice et Bob actifs ; Eve a une date_depart postérieure à la dernière séance."""
        import_all(test_xlsx.parent, test_db_path)
        out = tmp_path / "dashboard.html"
        generate(test_db_path, out)
        data = _extract_compas_json(out.read_text(encoding="utf-8"))
        noms = {s["name"] for s in data["students"]}
        assert "Dupont Alice" in noms
        assert "Martin Bob" in noms

    def test_import_destructif_idempotent(self, test_xlsx, test_db_path):
        """Deux import successifs donnent exactement le même résultat."""
        import_all(test_xlsx.parent, test_db_path)
        conn = sqlite3.connect(test_db_path)
        count1 = conn.execute("SELECT COUNT(*) FROM releves").fetchone()[0]
        conn.close()

        import_all(test_xlsx.parent, test_db_path)
        conn = sqlite3.connect(test_db_path)
        count2 = conn.execute("SELECT COUNT(*) FROM releves").fetchone()[0]
        conn.close()

        assert count1 == count2


# ---------------------------------------------------------------------------
# Tests end-to-end : deux fichiers xlsx (multi-projets)
# ---------------------------------------------------------------------------


class TestPipelineMultiFichiers:
    @pytest.fixture
    def data_dir(self, tmp_path):
        d = tmp_path / "data"
        d.mkdir()
        make_workbook().save(d / "projet01.xlsx")
        _make_second_workbook().save(d / "projet02.xlsx")
        return d

    def test_deux_projets_inseres(self, data_dir, tmp_path):
        db_path = tmp_path / "compas.db"
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            count = conn.execute("SELECT COUNT(*) FROM projets").fetchone()[0]
        finally:
            conn.close()
        assert count == 2

    def test_etudiant_partage_une_seule_entree(self, data_dir, tmp_path):
        """Dupont Alice présente dans les 2 xlsx → une seule ligne dans etudiants."""
        db_path = tmp_path / "compas.db"
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            count = conn.execute(
                "SELECT COUNT(*) FROM etudiants WHERE nom='Dupont Alice'"
            ).fetchone()[0]
        finally:
            conn.close()
        assert count == 1

    def test_quatre_etudiants_distincts(self, data_dir, tmp_path):
        """3 de projet01 + 1 unique de projet02 = 4 étudiants distincts."""
        db_path = tmp_path / "compas.db"
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            count = conn.execute("SELECT COUNT(*) FROM etudiants").fetchone()[0]
        finally:
            conn.close()
        assert count == 4

    def test_releves_cumules_des_deux_projets(self, data_dir, tmp_path):
        """projet01 : 4 relevés ; projet02 : 2 relevés → total = 6."""
        db_path = tmp_path / "compas.db"
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            count = conn.execute("SELECT COUNT(*) FROM releves").fetchone()[0]
        finally:
            conn.close()
        assert count == 6

    def test_warning_plusieurs_projets(self, data_dir, tmp_path, caplog):
        """La génération du dashboard avertit si plusieurs projets sont présents."""
        import logging

        db_path = tmp_path / "compas.db"
        out = tmp_path / "dashboard.html"
        import_all(data_dir, db_path)
        with caplog.at_level(logging.WARNING, logger="compas.dashboard"):
            generate(db_path, out)
        assert "Plusieurs projets" in caplog.text
