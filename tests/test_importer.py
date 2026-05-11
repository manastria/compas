"""Tests du module compas.importer."""

import logging
import sqlite3
from datetime import date, datetime

import openpyxl
import pytest

from compas.importer import (
    _parse_config,
    _parse_date,
    _parse_score,
    _parse_seance,
    _should_ignore_sheet,
    import_all,
)
from tests.conftest import make_workbook


# ---------------------------------------------------------------------------
# _should_ignore_sheet
# ---------------------------------------------------------------------------


class TestShouldIgnoreSheet:
    def test_config_exact(self):
        assert _should_ignore_sheet("Config") is True

    def test_config_wrong_case_not_ignored(self):
        assert _should_ignore_sheet("config") is False
        assert _should_ignore_sheet("CONFIG") is False

    def test_modele_sans_accent(self):
        assert _should_ignore_sheet("Modele") is True

    def test_modele_avec_accent(self):
        assert _should_ignore_sheet("Modèle") is True

    def test_modele_majuscules(self):
        assert _should_ignore_sheet("MODELE") is True

    def test_tmp_prefix_minuscule(self):
        assert _should_ignore_sheet("tmp-brouillon") is True

    def test_tmp_prefix_majuscule(self):
        assert _should_ignore_sheet("TMP-TEST") is True

    def test_tmp_prefix_mixte(self):
        assert _should_ignore_sheet("Tmp-Foo") is True

    def test_seance_normale(self):
        assert _should_ignore_sheet("S1") is False
        assert _should_ignore_sheet("2026-03-27") is False
        assert _should_ignore_sheet("Séance 4") is False


# ---------------------------------------------------------------------------
# _parse_date
# ---------------------------------------------------------------------------


class TestParseDate:
    def test_none_returns_none(self):
        assert _parse_date(None) is None

    def test_empty_string_returns_none(self):
        assert _parse_date("") is None
        assert _parse_date("   ") is None

    def test_datetime_object(self):
        assert _parse_date(datetime(2026, 3, 27, 10, 0)) == "2026-03-27"

    def test_date_object(self):
        assert _parse_date(date(2026, 3, 27)) == "2026-03-27"

    def test_text_dmy(self):
        assert _parse_date("27/03/2026") == "2026-03-27"

    def test_text_iso(self):
        assert _parse_date("2026-03-27") == "2026-03-27"

    def test_invalid_emits_warning(self, caplog):
        with caplog.at_level(logging.WARNING, logger="compas.importer"):
            result = _parse_date("pas-une-date")
        assert result is None
        assert "non reconnu" in caplog.text


# ---------------------------------------------------------------------------
# _parse_score
# ---------------------------------------------------------------------------


class TestParseScore:
    def test_none_returns_none(self):
        assert _parse_score(None) is None

    def test_empty_string_returns_none(self):
        assert _parse_score("") is None
        assert _parse_score("   ") is None

    @pytest.mark.parametrize("v", [-2, -1, 0, 1, 2])
    def test_valid_integers(self, v):
        assert _parse_score(v) == v

    def test_valid_string_integer(self):
        assert _parse_score("1") == 1
        assert _parse_score("-2") == -2

    def test_out_of_range_returns_none_with_warning(self, caplog):
        with caplog.at_level(logging.WARNING, logger="compas.importer"):
            result = _parse_score(3)
        assert result is None
        assert "hors plage" in caplog.text

    def test_non_integer_returns_none_with_warning(self, caplog):
        with caplog.at_level(logging.WARNING, logger="compas.importer"):
            result = _parse_score("abc")
        assert result is None
        assert "non entier" in caplog.text


# ---------------------------------------------------------------------------
# _parse_config
# ---------------------------------------------------------------------------


class TestParseConfig:
    def test_projet_et_groupe(self):
        wb = make_workbook()
        nom_projet, groupe, _ = _parse_config(wb["Config"])
        assert nom_projet == "Infrastructure réseau PME"
        assert groupe == "TP1"

    def test_nombre_etudiants(self):
        wb = make_workbook()
        _, _, etudiants = _parse_config(wb["Config"])
        assert len(etudiants) == 3

    def test_etudiant_non_anonyme(self):
        wb = make_workbook()
        _, _, etudiants = _parse_config(wb["Config"])
        alice = etudiants[0]
        assert alice["nom"] == "Dupont Alice"
        assert alice["ine"] == "233303259HE"
        assert alice["anonyme"] == 0
        assert alice["pseudo"] is None
        assert alice["date_depart"] is None

    def test_etudiant_anonyme_avec_pseudo(self):
        wb = make_workbook()
        _, _, etudiants = _parse_config(wb["Config"])
        bob = etudiants[1]
        assert bob["ine"] == "070288524AE"
        assert bob["anonyme"] == 1
        assert bob["pseudo"] == "BobM"

    def test_etudiant_ine_absent(self):
        wb = make_workbook()
        _, _, etudiants = _parse_config(wb["Config"])
        eve = etudiants[2]
        assert eve["ine"] is None

    def test_etudiant_avec_date_depart(self):
        wb = make_workbook()
        _, _, etudiants = _parse_config(wb["Config"])
        eve = etudiants[2]
        assert eve["date_depart"] == "2026-02-15"

    def test_projet_manquant_leve_erreur(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Config"
        # B1 est vide → ValueError attendue
        with pytest.raises(ValueError, match="Projet"):
            _parse_config(ws)


# ---------------------------------------------------------------------------
# _parse_seance
# ---------------------------------------------------------------------------

_KNOWN_NAMES = {"Dupont Alice", "Martin Bob", "Leclerc Eve"}


class TestParseSeance:
    def test_metadata_complete(self):
        wb = make_workbook()
        meta, _ = _parse_seance(wb["S1"], _KNOWN_NAMES)
        assert meta["seance"] == 1
        assert meta["date"] == "2026-01-15"
        assert meta["heure_debut"] == "8h00"
        assert meta["heure_fin"] == "12h00"
        assert meta["enseignant"] == "Prof Martin"

    def test_nombre_de_releves(self):
        wb = make_workbook()
        _, releves = _parse_seance(wb["S1"], _KNOWN_NAMES)
        assert len(releves) == 3

    def test_scores_alice(self):
        wb = make_workbook()
        _, releves = _parse_seance(wb["S1"], _KNOWN_NAMES)
        alice = next(r for r in releves if r["nom"] == "Dupont Alice")
        assert alice["presence"] == "P"
        assert alice["autonomie"] == 1
        assert alice["rigueur"] == 2
        assert alice["communication"] == 0
        assert alice["engagement"] == 1
        assert alice["commentaire"] == "Bon travail"

    def test_score_null_non_observe(self):
        wb = make_workbook()
        _, releves = _parse_seance(wb["S1"], _KNOWN_NAMES)
        bob = next(r for r in releves if r["nom"] == "Martin Bob")
        assert bob["rigueur"] is None

    def test_retard(self):
        wb = make_workbook()
        _, releves = _parse_seance(wb["S1"], _KNOWN_NAMES)
        bob = next(r for r in releves if r["nom"] == "Martin Bob")
        assert bob["presence"] == "R:15"

    def test_absence(self):
        wb = make_workbook()
        _, releves = _parse_seance(wb["S1"], _KNOWN_NAMES)
        eve = next(r for r in releves if r["nom"] == "Leclerc Eve")
        assert eve["presence"] == "A"

    def test_presence_vide_devient_p(self):
        wb = make_workbook()
        _, releves = _parse_seance(wb["S2"], _KNOWN_NAMES)
        alice = next(r for r in releves if r["nom"] == "Dupont Alice")
        assert alice["presence"] == "P"

    def test_etudiant_inconnu_emet_warning(self, caplog):
        wb = make_workbook()
        ws = wb["S1"]
        ws.cell(row=9, column=1, value="Inconnu Tartempion")
        with caplog.at_level(logging.WARNING, logger="compas.importer"):
            _, releves = _parse_seance(ws, _KNOWN_NAMES)
        assert "Inconnu Tartempion" in caplog.text
        assert not any(r["nom"] == "Inconnu Tartempion" for r in releves)


# ---------------------------------------------------------------------------
# Tests d'intégration : import_all
# ---------------------------------------------------------------------------


class TestImportAll:
    def test_cree_la_base(self, test_xlsx, test_db_path):
        import_all(test_xlsx.parent, test_db_path)
        assert test_db_path.exists()

    def test_projet_insere(self, test_xlsx, test_db_path):
        import_all(test_xlsx.parent, test_db_path)
        conn = sqlite3.connect(test_db_path)
        try:
            projets = conn.execute("SELECT nom, groupe FROM projets").fetchall()
            assert projets == [("Infrastructure réseau PME", "TP1")]
        finally:
            conn.close()

    def test_etudiants_inseres(self, test_xlsx, test_db_path):
        import_all(test_xlsx.parent, test_db_path)
        conn = sqlite3.connect(test_db_path)
        try:
            noms = {row[0] for row in conn.execute("SELECT nom FROM etudiants").fetchall()}
            assert noms == {"Dupont Alice", "Martin Bob", "Leclerc Eve"}
        finally:
            conn.close()

    def test_nombre_de_releves(self, test_xlsx, test_db_path):
        # S1 : Alice + Bob + Eve = 3  |  S2 : Alice = 1  →  total = 4
        import_all(test_xlsx.parent, test_db_path)
        conn = sqlite3.connect(test_db_path)
        try:
            count = conn.execute("SELECT COUNT(*) FROM releves").fetchone()[0]
            assert count == 4
        finally:
            conn.close()

    def test_scores_alice_seance1(self, test_xlsx, test_db_path):
        import_all(test_xlsx.parent, test_db_path)
        conn = sqlite3.connect(test_db_path)
        try:
            row = conn.execute(
                """SELECT r.autonomie, r.rigueur, r.communication, r.engagement, r.presence
                   FROM releves r
                   JOIN etudiants e ON r.etudiant_id = e.id
                   WHERE e.nom = 'Dupont Alice' AND r.seance = 1"""
            ).fetchone()
            assert row == (1, 2, 0, 1, "P")
        finally:
            conn.close()

    def test_rigueur_null_bob_seance1(self, test_xlsx, test_db_path):
        import_all(test_xlsx.parent, test_db_path)
        conn = sqlite3.connect(test_db_path)
        try:
            row = conn.execute(
                """SELECT r.rigueur FROM releves r
                   JOIN etudiants e ON r.etudiant_id = e.id
                   WHERE e.nom = 'Martin Bob' AND r.seance = 1"""
            ).fetchone()
            assert row[0] is None
        finally:
            conn.close()

    def test_feuilles_ignorees(self, tmp_path):
        """Les feuilles Config, Modèle et tmp-* ne créent aucun relevé."""
        data_dir = tmp_path / "data"
        data_dir.mkdir()
        wb = openpyxl.Workbook()
        config = wb.active
        config.title = "Config"
        config["A1"] = "Projet"
        config["B1"] = "Test"
        config["A2"] = "Groupe"
        config["B2"] = "G1"
        config.cell(row=4, column=1, value="Nom")
        config.cell(row=5, column=1, value="Dupont Alice")
        config.cell(row=5, column=3, value="non")  # col C = Anonyme
        wb.create_sheet("Modèle")
        wb.create_sheet("tmp-brouillon")
        wb.save(data_dir / "test.xlsx")

        db_path = tmp_path / "compas.db"
        import_all(data_dir, db_path)

        conn = sqlite3.connect(db_path)
        try:
            assert conn.execute("SELECT COUNT(*) FROM releves").fetchone()[0] == 0
        finally:
            conn.close()

    def test_deduplication_meme_etudiant_deux_fichiers(self, tmp_path):
        """Un étudiant identique dans 2 xlsx → une seule ligne dans etudiants."""
        data_dir = tmp_path / "data"
        data_dir.mkdir()

        for i, pseudo in enumerate(["PseudoA", "PseudoB"], start=1):
            wb = openpyxl.Workbook()
            config = wb.active
            config.title = "Config"
            config["A1"] = "Projet"
            config["B1"] = f"Projet {i}"
            config["A2"] = "Groupe"
            config["B2"] = "TP1"
            config.cell(row=4, column=1, value="Nom")
            config.cell(row=5, column=1, value="Dupont Alice")
            config.cell(row=5, column=3, value="oui")   # col C = Anonyme
            config.cell(row=5, column=4, value=pseudo)  # col D = Pseudo
            # tri lexicographique → projet02 lu en dernier
            wb.save(data_dir / f"projet{i:02d}.xlsx")

        db_path = tmp_path / "compas.db"
        import_all(data_dir, db_path)

        conn = sqlite3.connect(db_path)
        try:
            count = conn.execute(
                "SELECT COUNT(*) FROM etudiants WHERE nom = 'Dupont Alice'"
            ).fetchone()[0]
            assert count == 1
            # Le fichier projet02.xlsx (PseudoB) est lu en dernier → prévaut
            pseudo_result = conn.execute(
                "SELECT pseudo FROM etudiants WHERE nom = 'Dupont Alice'"
            ).fetchone()[0]
            assert pseudo_result == "PseudoB"
        finally:
            conn.close()

    def test_avertissement_aucun_fichier(self, tmp_path, caplog):
        data_dir = tmp_path / "empty"
        data_dir.mkdir()
        with caplog.at_level(logging.WARNING, logger="compas.importer"):
            import_all(data_dir, tmp_path / "compas.db")
        assert "Aucun fichier xlsx" in caplog.text

    def test_config_manquante_leve_erreur(self, tmp_path):
        data_dir = tmp_path / "data"
        data_dir.mkdir()
        wb = openpyxl.Workbook()
        wb.active.title = "Séance1"  # pas de feuille Config
        wb.save(data_dir / "mauvais.xlsx")

        with pytest.raises(ValueError, match="Config"):
            import_all(data_dir, tmp_path / "compas.db")


# ---------------------------------------------------------------------------
# Stratégie last-file-wins : tous les champs en conflit
# ---------------------------------------------------------------------------


def _make_two_configs(
    tmp_path,
    *,
    ine1=None,
    ine2=None,
    anonyme1="non",
    anonyme2="non",
    pseudo1=None,
    pseudo2=None,
    date_depart1=None,
    date_depart2=None,
) -> tuple:
    """Crée deux xlsx avec le même étudiant, champs potentiellement différents.

    Retourne (data_dir, db_path).
    """
    data_dir = tmp_path / "data"
    data_dir.mkdir()

    for i, (ine, anon, pseudo, depart) in enumerate(
        [
            (ine1, anonyme1, pseudo1, date_depart1),
            (ine2, anonyme2, pseudo2, date_depart2),
        ],
        start=1,
    ):
        wb = openpyxl.Workbook()
        config = wb.active
        config.title = "Config"
        config["A1"] = "Projet"
        config["B1"] = f"Projet {i}"
        config["A2"] = "Groupe"
        config["B2"] = "TP1"
        config.cell(row=4, column=1, value="Nom")
        config.cell(row=5, column=1, value="Dupont Alice")
        if ine:
            config.cell(row=5, column=2, value=ine)
        config.cell(row=5, column=3, value=anon)
        if pseudo:
            config.cell(row=5, column=4, value=pseudo)
        if depart:
            config.cell(row=5, column=5, value=depart)
        wb.save(data_dir / f"projet{i:02d}.xlsx")

    return data_dir, tmp_path / "compas.db"


class TestLastFileWins:
    """Vérifie la stratégie « dernier fichier lu prévaut » pour tous les champs en conflit."""

    def test_ine_last_wins(self, tmp_path):
        data_dir, db_path = _make_two_configs(tmp_path, ine1="INE_A", ine2="INE_B")
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            ine = conn.execute(
                "SELECT ine FROM etudiants WHERE nom='Dupont Alice'"
            ).fetchone()[0]
        finally:
            conn.close()
        assert ine == "INE_B"

    def test_anonyme_last_wins(self, tmp_path):
        data_dir, db_path = _make_two_configs(
            tmp_path, anonyme1="non", anonyme2="oui", pseudo2="AliceX"
        )
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            row = conn.execute(
                "SELECT anonyme, pseudo FROM etudiants WHERE nom='Dupont Alice'"
            ).fetchone()
        finally:
            conn.close()
        assert row[0] == 1
        assert row[1] == "AliceX"

    def test_date_depart_last_wins(self, tmp_path):
        data_dir, db_path = _make_two_configs(
            tmp_path, date_depart1="2026-01-01", date_depart2="2026-06-15"
        )
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            dd = conn.execute(
                "SELECT date_depart FROM etudiants WHERE nom='Dupont Alice'"
            ).fetchone()[0]
        finally:
            conn.close()
        assert dd == "2026-06-15"

    def test_ine_second_file_wins_even_if_null(self, tmp_path):
        """Le second fichier écrase l'INE même si le second n'en a pas (NULL prévaut)."""
        data_dir, db_path = _make_two_configs(tmp_path, ine1="INE_A", ine2=None)
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            ine = conn.execute(
                "SELECT ine FROM etudiants WHERE nom='Dupont Alice'"
            ).fetchone()[0]
        finally:
            conn.close()
        assert ine is None

    def test_une_seule_entree_par_etudiant(self, tmp_path):
        """Quel que soit le nombre de fichiers, l'étudiant n'est inséré qu'une fois."""
        data_dir, db_path = _make_two_configs(
            tmp_path, ine1="INE_A", ine2="INE_A", pseudo1="A1", pseudo2="A2"
        )
        import_all(data_dir, db_path)
        conn = sqlite3.connect(db_path)
        try:
            count = conn.execute(
                "SELECT COUNT(*) FROM etudiants WHERE nom='Dupont Alice'"
            ).fetchone()[0]
        finally:
            conn.close()
        assert count == 1
