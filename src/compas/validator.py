"""Validation de la conformité des fichiers XLSX pour COMPAS."""

import re
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from enum import Enum
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class Severity(Enum):
    ERROR = "ERREUR"
    WARNING = "AVERT."


@dataclass
class Issue:
    severity: Severity
    sheet: str
    location: str  # ex : "B1", "ligne 5 col B"
    message: str

    def __str__(self) -> str:
        coord = f"{self.sheet}:{self.location}" if self.location else self.sheet
        return f"{self.severity.value:<7}  {coord:<28}  {self.message}"


# ---------------------------------------------------------------------------
# Patterns de validation
# ---------------------------------------------------------------------------

_RE_INE = re.compile(r"^\d{9}[A-Z]{2}$")
_RE_HEURE = r"\d{1,2}h\d{2}"
_RE_MOTIF = r"[^\s:,]+"
_RE_PLAGE = rf"(?:{_RE_HEURE}-{_RE_HEURE}|H\d+-H\d+)"

_PRESENCE_TOKENS = [
    re.compile(r"^P$"),
    re.compile(r"^N$"),
    re.compile(rf"^A(?::{_RE_PLAGE}(?::{_RE_MOTIF})?|:{_RE_MOTIF})?$"),
    re.compile(rf"^R:(?:\d+|{_RE_HEURE})(?::{_RE_MOTIF})?$"),
    re.compile(rf"^RR:\d+(?::{_RE_MOTIF})?$"),
    re.compile(rf"^D:{_RE_HEURE}(?::{_RE_MOTIF})?$"),
]

_SCORE_COLS = {3: "Autonomie", 4: "Rigueur", 5: "Communication", 6: "Engagement"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _ignore_sheet(name: str) -> bool:
    """Même règle que l'importeur : Config, Modèle, tmp-*."""
    if name == "Config":
        return True
    normalized = unicodedata.normalize("NFD", name).encode("ascii", "ignore").decode().lower()
    if normalized == "modele":
        return True
    return name.lower().startswith("tmp-")


def _parse_date_ok(value: object) -> bool:
    """Retourne True si value est une date parseable."""
    if value is None:
        return False
    if isinstance(value, (datetime, date)):
        return True
    text = str(value).strip()
    if not text:
        return False
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            pass
    return False


def _invalid_presence_tokens(value: str) -> list[str]:
    """Retourne les tokens de la chaîne de présence qui ne correspondent à aucun format connu."""
    if not value or value.strip() in ("", "P"):
        return []
    return [
        tok.strip()
        for tok in value.split(",")
        if tok.strip() and not any(pat.match(tok.strip()) for pat in _PRESENCE_TOKENS)
    ]


# ---------------------------------------------------------------------------
# Validation Config
# ---------------------------------------------------------------------------


def _validate_config(ws: Worksheet) -> tuple[list[Issue], list[str]]:
    """Valide la feuille Config.

    Returns:
        (issues, noms_etudiants)
    """
    issues: list[Issue] = []
    noms: list[str] = []

    if not ws["B1"].value or not str(ws["B1"].value).strip():
        issues.append(Issue(Severity.ERROR, "Config", "B1", "Nom du projet manquant"))
        return issues, noms

    if not ws["B2"].value or not str(ws["B2"].value).strip():
        issues.append(Issue(Severity.WARNING, "Config", "B2", "Nom du groupe manquant"))

    header_a4 = ws.cell(row=4, column=1).value
    if header_a4 is None or str(header_a4).strip().lower() != "nom":
        issues.append(Issue(
            Severity.WARNING, "Config", "A4",
            f"En-tête 'Nom' attendu en A4, trouvé : {header_a4!r}",
        ))

    row = 5
    while True:
        nom_val = ws.cell(row=row, column=1).value
        if nom_val is None or str(nom_val).strip() == "":
            break
        nom = str(nom_val).strip()
        noms.append(nom)

        # INE (col B) : format si présent
        ine_val = ws.cell(row=row, column=2).value
        if ine_val is not None and str(ine_val).strip():
            ine = str(ine_val).strip()
            if not _RE_INE.match(ine):
                issues.append(Issue(
                    Severity.WARNING, "Config", f"ligne {row} col B",
                    f"INE '{ine}' : format attendu 9 chiffres + 2 majuscules (ex : 233303259HE)",
                ))

        # Anonyme (col C) : "oui" / "non"
        anon_val = ws.cell(row=row, column=3).value
        anon_str = str(anon_val).strip().lower() if anon_val is not None else ""
        if anon_str not in ("oui", "non", ""):
            issues.append(Issue(
                Severity.ERROR, "Config", f"ligne {row} col C",
                f"Valeur 'Anonyme' invalide : {anon_val!r} (attendu : 'oui' ou 'non')",
            ))
        elif anon_str == "oui":
            pseudo_val = ws.cell(row=row, column=4).value
            if not pseudo_val or not str(pseudo_val).strip():
                issues.append(Issue(
                    Severity.WARNING, "Config", f"ligne {row} col D",
                    f"Étudiant '{nom}' marqué anonyme sans pseudo (col D vide)",
                ))

        # Date de départ (col E)
        date_val = ws.cell(row=row, column=5).value
        if date_val is not None and str(date_val).strip():
            if not _parse_date_ok(date_val):
                issues.append(Issue(
                    Severity.ERROR, "Config", f"ligne {row} col E",
                    f"Date de départ de '{nom}' non reconnue : {date_val!r}",
                ))

        row += 1

    if not noms:
        issues.append(Issue(Severity.WARNING, "Config", "ligne 5+", "Aucun étudiant trouvé"))

    return issues, noms


# ---------------------------------------------------------------------------
# Validation feuille de séance
# ---------------------------------------------------------------------------


def _validate_seance(
    ws: Worksheet, known_names: set[str]
) -> tuple[list[Issue], Optional[int]]:
    """Valide une feuille de séance.

    Returns:
        (issues, seance_num) — seance_num est None si illisible.
    """
    issues: list[Issue] = []
    sheet = ws.title
    seance_num: Optional[int] = None

    # B2 : numéro de séance
    b2 = ws["B2"].value
    if b2 is None:
        issues.append(Issue(Severity.ERROR, sheet, "B2", "Numéro de séance manquant"))
    else:
        try:
            seance_num = int(b2)
        except (ValueError, TypeError):
            issues.append(Issue(
                Severity.ERROR, sheet, "B2", f"Numéro de séance non entier : {b2!r}"
            ))

    # D2 : date
    d2 = ws["D2"].value
    if d2 is None or (isinstance(d2, str) and not d2.strip()):
        issues.append(Issue(Severity.ERROR, sheet, "D2", "Date de séance manquante"))
    elif not _parse_date_ok(d2):
        issues.append(Issue(
            Severity.ERROR, sheet, "D2",
            f"Format de date non reconnu : {d2!r} (attendu DD/MM/YYYY ou YYYY-MM-DD)",
        ))

    # F2, H2, J2 : métadonnées optionnelles mais attendues
    if not ws["F2"].value:
        issues.append(Issue(Severity.WARNING, sheet, "F2", "Heure de début manquante"))
    if not ws["H2"].value:
        issues.append(Issue(Severity.WARNING, sheet, "H2", "Nom de l'enseignant manquant"))
    if not ws["J2"].value:
        issues.append(Issue(Severity.WARNING, sheet, "J2", "Heure de fin manquante"))

    # Lignes 6+ : relevés
    row = 6
    while True:
        nom_val = ws.cell(row=row, column=1).value
        if nom_val is None or str(nom_val).strip() == "":
            break
        nom = str(nom_val).strip()

        if known_names and nom not in known_names:
            issues.append(Issue(
                Severity.WARNING, sheet, f"ligne {row} col A",
                f"Étudiant '{nom}' absent de la feuille Config",
            ))

        # Présence (col B)
        pres_val = ws.cell(row=row, column=2).value
        if pres_val is not None:
            pres = str(pres_val).strip()
            for bad_tok in _invalid_presence_tokens(pres):
                issues.append(Issue(
                    Severity.WARNING, sheet, f"ligne {row} col B",
                    f"Token de présence non reconnu : '{bad_tok}' (valeur : '{pres}')",
                ))

        # Scores cols C–F
        for col, label in _SCORE_COLS.items():
            val = ws.cell(row=row, column=col).value
            if val is not None:
                try:
                    score = int(val)
                    if not -2 <= score <= 2:
                        issues.append(Issue(
                            Severity.ERROR, sheet, f"ligne {row} col {chr(64 + col)}",
                            f"{label} hors plage [-2, +2] pour '{nom}' : {val}",
                        ))
                except (ValueError, TypeError):
                    issues.append(Issue(
                        Severity.ERROR, sheet, f"ligne {row} col {chr(64 + col)}",
                        f"{label} non entier pour '{nom}' : {val!r}",
                    ))

        row += 1

    return issues, seance_num


# ---------------------------------------------------------------------------
# Point d'entrée public
# ---------------------------------------------------------------------------


def validate_xlsx(path: Path) -> list[Issue]:
    """Valide un fichier XLSX COMPAS et retourne la liste des problèmes détectés.

    Args:
        path: Chemin vers le fichier .xlsx à valider.

    Returns:
        Liste d'Issue (erreurs et avertissements). Liste vide = fichier conforme.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:
        return [Issue(Severity.ERROR, "(fichier)", "", f"Impossible d'ouvrir le fichier : {exc}")]

    issues: list[Issue] = []

    if "Config" not in wb.sheetnames:
        issues.append(Issue(Severity.ERROR, "(fichier)", "", "Feuille 'Config' absente"))
        return issues

    config_issues, noms = _validate_config(wb["Config"])
    issues.extend(config_issues)
    known_names = set(noms)

    seance_nums: list[tuple[int, str]] = []
    for sheet_name in wb.sheetnames:
        if _ignore_sheet(sheet_name):
            continue
        ws = wb[sheet_name]
        sheet_issues, seance_num = _validate_seance(ws, known_names)
        issues.extend(sheet_issues)
        if seance_num is not None:
            seance_nums.append((seance_num, sheet_name))

    # Doublons de numéros de séance
    seen: dict[int, str] = {}
    for num, name in seance_nums:
        if num in seen:
            issues.append(Issue(
                Severity.WARNING, name, "B2",
                f"Numéro de séance {num} déjà utilisé dans '{seen[num]}'",
            ))
        else:
            seen[num] = name

    if not seance_nums:
        issues.append(Issue(
            Severity.WARNING, "(fichier)", "",
            "Aucune feuille de séance trouvée (toutes absentes ou ignorées)",
        ))

    return issues
