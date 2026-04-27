"""Lecture des fichiers xlsx et insertion dans SQLite."""

import logging
import sqlite3
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

_INIT_SQL = """
DROP TABLE IF EXISTS releves;
DROP TABLE IF EXISTS projets;
DROP TABLE IF EXISTS etudiants;

CREATE TABLE etudiants (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nom TEXT NOT NULL,
    groupe TEXT,
    ine TEXT UNIQUE,
    anonyme INTEGER NOT NULL DEFAULT 0,
    pseudo TEXT,
    date_depart TEXT
);

CREATE TABLE projets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nom TEXT NOT NULL,
    groupe TEXT
);

CREATE TABLE releves (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    etudiant_id INTEGER NOT NULL REFERENCES etudiants(id),
    projet_id INTEGER NOT NULL REFERENCES projets(id),
    seance INTEGER NOT NULL,
    date TEXT NOT NULL,
    heure_debut TEXT,
    heure_fin TEXT,
    enseignant TEXT,
    presence TEXT,
    autonomie INTEGER,
    rigueur INTEGER,
    communication INTEGER,
    engagement INTEGER,
    commentaire TEXT,
    UNIQUE(etudiant_id, projet_id, seance)
);
"""


def _should_ignore_sheet(name: str) -> bool:
    """Retourne True si la feuille doit être ignorée lors de l'import.

    Règles :
    - « Config » : correspondance exacte (sensible à la casse)
    - « Modele » / « Modèle » : insensible aux accents
    - Préfixe « tmp- » : insensible à la casse
    """
    if name == "Config":
        return True
    normalized = unicodedata.normalize("NFD", name).encode("ascii", "ignore").decode().lower()
    if normalized == "modele":
        return True
    if name.lower().startswith("tmp-"):
        return True
    return False


def _parse_date(value: object) -> Optional[str]:
    """Parse une date Excel ou texte en format ISO YYYY-MM-DD.

    Args:
        value: Valeur brute de cellule (datetime, date, str, None).

    Returns:
        Chaîne ISO YYYY-MM-DD ou None si absent/invalide.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    if not text:
        return None
    # Format DD/MM/YYYY
    try:
        return datetime.strptime(text, "%d/%m/%Y").strftime("%Y-%m-%d")
    except ValueError:
        pass
    # Format ISO YYYY-MM-DD (déjà correct)
    try:
        datetime.strptime(text, "%Y-%m-%d")
        return text
    except ValueError:
        pass
    logger.warning("Format de date non reconnu : %r", value)
    return None


def _parse_score(value: object) -> Optional[int]:
    """Parse un score entier compris entre -2 et +2.

    Args:
        value: Valeur brute de cellule.

    Returns:
        Entier ou None si absent ou invalide.
    """
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        score = int(value)
    except (ValueError, TypeError):
        logger.warning("Score non entier ignoré : %r", value)
        return None
    if not -2 <= score <= 2:
        logger.warning("Score hors plage [-2, 2] ignoré : %r", value)
        return None
    return score


def _parse_config(ws: Worksheet) -> tuple[str, str, list[dict]]:
    """Lit la feuille Config et retourne le projet, le groupe et la liste des étudiants.

    Args:
        ws: Feuille openpyxl nommée « Config ».

    Returns:
        Tuple (nom_projet, groupe, etudiants) où etudiants est une liste de dicts
        avec les clés : nom, ine, anonyme (0/1), pseudo, date_depart.
        Colonnes : A=Nom, B=INE, C=Anonyme, D=Pseudo, E=Date de départ.

    Raises:
        ValueError: Si le champ « Projet » (B1) est absent.
    """
    nom_projet = ws["B1"].value
    if not nom_projet:
        raise ValueError("Champ 'Projet' (B1) manquant dans la feuille Config")
    nom_projet = str(nom_projet).strip()

    groupe = ws["B2"].value or ""
    groupe = str(groupe).strip()

    etudiants: list[dict] = []
    row = 5  # ligne 4 = en-têtes, données à partir de la ligne 5
    while True:
        nom_cell = ws.cell(row=row, column=1).value
        if nom_cell is None or str(nom_cell).strip() == "":
            break
        nom = str(nom_cell).strip()

        ine_raw = ws.cell(row=row, column=2).value
        ine: Optional[str] = str(ine_raw).strip() if ine_raw else None
        if ine == "":
            ine = None

        anonyme_raw = ws.cell(row=row, column=3).value
        anonyme = 1 if (anonyme_raw and str(anonyme_raw).strip().lower() == "oui") else 0

        pseudo_raw = ws.cell(row=row, column=4).value
        pseudo: Optional[str] = str(pseudo_raw).strip() if pseudo_raw else None
        if pseudo == "":
            pseudo = None

        date_depart = _parse_date(ws.cell(row=row, column=5).value)

        etudiants.append(
            {
                "nom": nom,
                "ine": ine,
                "anonyme": anonyme,
                "pseudo": pseudo,
                "date_depart": date_depart,
            }
        )
        row += 1

    return nom_projet, groupe, etudiants


def _parse_seance(ws: Worksheet, known_names: set[str]) -> tuple[dict, list[dict]]:
    """Lit une feuille de séance et retourne ses métadonnées et les relevés.

    Args:
        ws: Feuille openpyxl de séance.
        known_names: Ensemble des noms d'étudiants valides (issus de la feuille Config).

    Returns:
        Tuple (metadata, releves).
        - metadata : dict avec seance, date, heure_debut, enseignant.
        - releves : liste de dicts par étudiant avec nom, presence, autonomie,
          rigueur, communication, engagement, commentaire.
    """
    # Numéro de séance (B2)
    seance_num: Optional[int] = None
    seance_raw = ws["B2"].value
    if seance_raw is not None:
        try:
            seance_num = int(seance_raw)
        except (ValueError, TypeError):
            logger.warning(
                "Numéro de séance invalide dans la feuille '%s' : %r", ws.title, seance_raw
            )

    date_seance = _parse_date(ws["D2"].value)
    if date_seance is None:
        logger.warning("Date manquante ou invalide dans la feuille '%s'", ws.title)

    heure_debut_raw = ws["F2"].value
    heure_debut: Optional[str] = str(heure_debut_raw).strip() if heure_debut_raw else None

    enseignant_raw = ws["H2"].value
    enseignant: Optional[str] = str(enseignant_raw).strip() if enseignant_raw else None

    heure_fin_raw = ws["J2"].value
    heure_fin: Optional[str] = str(heure_fin_raw).strip() if heure_fin_raw else None

    metadata = {
        "seance": seance_num,
        "date": date_seance,
        "heure_debut": heure_debut,
        "heure_fin": heure_fin,
        "enseignant": enseignant,
    }

    releves: list[dict] = []
    row = 6  # ligne 4 = en-têtes, ligne 5 = symboles, données à partir de la ligne 6
    while True:
        nom_cell = ws.cell(row=row, column=1).value
        if nom_cell is None or str(nom_cell).strip() == "":
            break
        nom = str(nom_cell).strip()

        if nom not in known_names:
            logger.warning(
                "Étudiant '%s' inconnu dans la feuille '%s' (absent de Config)", nom, ws.title
            )
            row += 1
            continue

        presence_raw = ws.cell(row=row, column=2).value
        presence: str = str(presence_raw).strip() if presence_raw else "P"
        if not presence:
            presence = "P"

        autonomie = _parse_score(ws.cell(row=row, column=3).value)
        rigueur = _parse_score(ws.cell(row=row, column=4).value)
        communication = _parse_score(ws.cell(row=row, column=5).value)
        engagement = _parse_score(ws.cell(row=row, column=6).value)

        commentaire_raw = ws.cell(row=row, column=7).value
        commentaire: Optional[str] = str(commentaire_raw).strip() if commentaire_raw else None
        if commentaire == "":
            commentaire = None

        releves.append(
            {
                "nom": nom,
                "presence": presence,
                "autonomie": autonomie,
                "rigueur": rigueur,
                "communication": communication,
                "engagement": engagement,
                "commentaire": commentaire,
            }
        )
        row += 1

    return metadata, releves


def _init_db(conn: sqlite3.Connection) -> None:
    """Supprime et recrée les tables SQLite (import destructif)."""
    conn.executescript(_INIT_SQL)


def _upsert_etudiant(
    conn: sqlite3.Connection,
    nom: str,
    groupe: str,
    ine: Optional[str],
    anonyme: int,
    pseudo: Optional[str],
    date_depart: Optional[str],
    cache: dict[str, int],
) -> int:
    """Insère ou met à jour un étudiant et retourne son id.

    Si l'étudiant est déjà dans le cache (vu dans un fichier précédent),
    ses champs sont mis à jour avec les valeurs du fichier courant (dernier lu prévaut).

    Args:
        conn: Connexion SQLite active.
        nom: Nom complet de l'étudiant.
        groupe: Groupe associé au projet courant.
        ine: Identifiant national étudiant (peut être None).
        anonyme: 1 si anonyme, 0 sinon.
        pseudo: Pseudo pour le dashboard (peut être None).
        date_depart: Date de départ ISO ou None.
        cache: Dictionnaire partagé nom → id pour le dédoublonnage inter-fichiers.

    Returns:
        Identifiant SQLite de l'étudiant.
    """
    if nom in cache:
        etudiant_id = cache[nom]
        conn.execute(
            "UPDATE etudiants SET groupe=?, ine=?, anonyme=?, pseudo=?, date_depart=? WHERE id=?",
            (groupe, ine, anonyme, pseudo, date_depart, etudiant_id),
        )
        return etudiant_id

    cursor = conn.execute(
        "INSERT INTO etudiants (nom, groupe, ine, anonyme, pseudo, date_depart) VALUES (?, ?, ?, ?, ?, ?)",
        (nom, groupe, ine, anonyme, pseudo, date_depart),
    )
    etudiant_id = cursor.lastrowid
    cache[nom] = etudiant_id
    return etudiant_id


def import_xlsx(
    xlsx_path: Path,
    conn: sqlite3.Connection,
    etudiant_cache: dict[str, int],
) -> None:
    """Importe un fichier xlsx dans la base SQLite.

    Lit la feuille Config pour créer le projet et les étudiants, puis parcourt
    toutes les feuilles de séance pour insérer les relevés.

    Args:
        xlsx_path: Chemin vers le fichier .xlsx.
        conn: Connexion SQLite active (sans auto-commit).
        etudiant_cache: Cache partagé nom → id pour le dédoublonnage inter-fichiers.

    Raises:
        ValueError: Si le fichier est illisible ou si la feuille Config est absente.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as exc:
        raise ValueError(f"Impossible d'ouvrir {xlsx_path.name} : {exc}") from exc

    if "Config" not in wb.sheetnames:
        raise ValueError(f"Feuille 'Config' manquante dans {xlsx_path.name}")

    nom_projet, groupe, etudiants = _parse_config(wb["Config"])

    # Insérer le projet
    cursor = conn.execute(
        "INSERT INTO projets (nom, groupe) VALUES (?, ?)",
        (nom_projet, groupe),
    )
    projet_id = cursor.lastrowid

    # Insérer / mettre à jour les étudiants
    known_names: set[str] = set()
    for etudiant in etudiants:
        _upsert_etudiant(
            conn,
            nom=etudiant["nom"],
            groupe=groupe,
            ine=etudiant["ine"],
            anonyme=etudiant["anonyme"],
            pseudo=etudiant["pseudo"],
            date_depart=etudiant["date_depart"],
            cache=etudiant_cache,
        )
        known_names.add(etudiant["nom"])

    # Parcourir les feuilles de séance
    for sheet_name in wb.sheetnames:
        if _should_ignore_sheet(sheet_name):
            continue

        ws = wb[sheet_name]
        metadata, releves = _parse_seance(ws, known_names)

        if metadata["seance"] is None:
            logger.warning("Feuille '%s' ignorée : numéro de séance manquant", sheet_name)
            continue
        if metadata["date"] is None:
            logger.warning("Feuille '%s' ignorée : date manquante ou invalide", sheet_name)
            continue

        for releve in releves:
            etudiant_id = etudiant_cache[releve["nom"]]
            try:
                conn.execute(
                    """INSERT OR REPLACE INTO releves
                       (etudiant_id, projet_id, seance, date, heure_debut, heure_fin, enseignant,
                        presence, autonomie, rigueur, communication, engagement, commentaire)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        etudiant_id,
                        projet_id,
                        metadata["seance"],
                        metadata["date"],
                        metadata["heure_debut"],
                        metadata["heure_fin"],
                        metadata["enseignant"],
                        releve["presence"],
                        releve["autonomie"],
                        releve["rigueur"],
                        releve["communication"],
                        releve["engagement"],
                        releve["commentaire"],
                    ),
                )
            except sqlite3.IntegrityError as exc:
                logger.warning(
                    "Doublon ignoré — étudiant '%s', projet '%s', séance %s : %s",
                    releve["nom"],
                    nom_projet,
                    metadata["seance"],
                    exc,
                )


def import_all(data_dir: Path, db_path: Path) -> None:
    """Importe tous les fichiers xlsx d'un répertoire dans une base SQLite.

    L'import est destructif : la base est reconstruite intégralement à chaque appel.

    Args:
        data_dir: Répertoire contenant les fichiers .xlsx.
        db_path: Chemin de la base SQLite à créer ou recréer.

    Raises:
        ValueError: Si un fichier xlsx est illisible ou mal formé (Config manquante).
    """
    db_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(db_path)
    try:
        _init_db(conn)
        etudiant_cache: dict[str, int] = {}

        xlsx_files = sorted(data_dir.glob("*.xlsx"))
        if not xlsx_files:
            logger.warning("Aucun fichier xlsx trouvé dans %s", data_dir)
            conn.commit()
            return

        for xlsx_path in xlsx_files:
            logger.info("Import : %s", xlsx_path.name)
            import_xlsx(xlsx_path, conn, etudiant_cache)

        conn.commit()
        logger.info(
            "Import terminé — %d étudiant(s), %d fichier(s)",
            len(etudiant_cache),
            len(xlsx_files),
        )
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
