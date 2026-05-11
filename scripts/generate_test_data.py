#!/usr/bin/env python3
"""Génère des données de test réalistes pour COMPAS (XLSX + CSV optionnel).

Usage:
    poetry run python scripts/generate_test_data.py
    poetry run python scripts/generate_test_data.py --out data/mon_projet.xlsx
    poetry run python scripts/generate_test_data.py --csv output/releves.csv
    poetry run python scripts/generate_test_data.py --seed 42 --students 12 --sessions 4
    poetry run python scripts/generate_test_data.py --fixture   # met à jour tests/fixtures/
"""

import argparse
import csv
import random
import string
from datetime import date, timedelta
from pathlib import Path

try:
    from faker import Faker
except ImportError:
    raise SystemExit(
        "Module 'faker' manquant. Installez-le : poetry add --group dev faker"
    )

import openpyxl
from openpyxl.styles import Alignment, Font

# ---------------------------------------------------------------------------
# Données de référence
# ---------------------------------------------------------------------------

PROJETS = [
    "Infrastructure réseau PME",
    "Administration système Linux",
    "Sécurité et pare-feu",
    "Virtualisation et conteneurs",
    "Supervision et monitoring",
    "Déploiement cloud AWS",
]

GROUPES = ["TP1", "TP2", "TP3", "G1", "G2"]

ENSEIGNANTS = ["M. Martin", "Mme Dubois", "M. Lambert", "Mme Perrot", "M. Rousseau"]

# (heure_debut, heure_fin)
CRENEAUX = [
    ("8h00", "10h00"),
    ("8h00", "12h00"),
    ("10h15", "12h15"),
    ("13h30", "15h30"),
    ("14h00", "18h00"),
]

MOTIFS_ABSENCE = ["medical", "transport", "famille", "stage", "sport"]

COMMENTAIRES_POS = [
    "Bon travail en équipe",
    "Très bonne participation",
    "Progrès notables",
    "Excellent niveau technique",
    "Travail soigné et rigoureux",
    "Prend des initiatives",
]

COMMENTAIRES_NEG = [
    "Manque d'initiative",
    "Doit relire les consignes",
    "Distrait en fin de séance",
    "Communication à améliorer",
    "Effort à fournir",
    "Travail en dessous du niveau attendu",
]

# ---------------------------------------------------------------------------
# Générateurs atomiques
# ---------------------------------------------------------------------------


def gen_ine() -> str:
    """Génère un INE fictif au format 9 chiffres + 2 majuscules (ex : 233303259HE)."""
    return "".join(random.choices(string.digits, k=9)) + "".join(
        random.choices(string.ascii_uppercase, k=2)
    )


def gen_presence(heure_debut: str) -> str:
    """Génère une valeur de présence selon la syntaxe COMPAS."""
    h_start = int(heure_debut.split("h")[0])
    r = random.random()

    if r < 0.65:
        return "P"

    if r < 0.75:
        # Retard en minutes
        minutes = random.choice([5, 10, 15, 20, 30])
        return f"R:{minutes}"

    if r < 0.80:
        # Arrivée à heure précise
        m = random.choice([10, 15, 20, 25, 30, 40, 45])
        return f"R:{h_start}h{m:02d}"

    if r < 0.88:
        # Absence
        if random.random() < 0.4:
            motif = random.choice(MOTIFS_ABSENCE)
            return f"A:{motif}"
        return "A"

    if r < 0.92:
        # Départ anticipé
        depart_h = h_start + random.randint(1, 2)
        depart_m = random.choice([0, 15, 30, 45])
        return f"D:{depart_h}h{depart_m:02d}"

    if r < 0.96:
        # Retard après récréation
        minutes = random.choice([5, 10, 15])
        return f"RR:{minutes}"

    if r < 0.98:
        # Retard début + retard récré
        r_min = random.choice([5, 10])
        rr_min = random.choice([5, 10])
        return f"R:{r_min},RR:{rr_min}"

    return "N"


def gen_score(base: float, variance: float, trend: float, seance_idx: int) -> int | None:
    """Génère un score -2..+2 selon le profil de l'étudiant.

    Args:
        base: Niveau de base de la compétence [-1.5, +1.5].
        variance: Amplitude de variation aléatoire [0, 1].
        trend: Évolution par séance [-0.2, +0.2].
        seance_idx: Index 0-based de la séance (pour appliquer la tendance).

    Returns:
        Entier -2..+2 ou None (non observé, ~15 % des cas).
    """
    if random.random() < 0.15:
        return None
    raw = base + trend * seance_idx + random.uniform(-variance, variance)
    return max(-2, min(2, round(raw)))


# ---------------------------------------------------------------------------
# Génération des entités
# ---------------------------------------------------------------------------


def generate_students(fake: "Faker", n: int) -> list[dict]:
    """Génère n étudiants avec profils de performance et métadonnées."""
    students: list[dict] = []
    seen_ines: set[str] = set()

    for _ in range(n):
        ine = gen_ine()
        while ine in seen_ines:
            ine = gen_ine()
        seen_ines.add(ine)

        nom = f"{fake.last_name().upper()} {fake.first_name()}"
        anonyme = random.random() < 0.15
        pseudo = fake.user_name() if anonyme else None

        date_depart: date | None = None
        if random.random() < 0.08:
            date_depart = date.today() - timedelta(days=random.randint(30, 90))

        students.append(
            {
                "nom": nom,
                "ine": ine,
                "anonyme": "oui" if anonyme else "non",
                "pseudo": pseudo,
                "date_depart": date_depart.strftime("%Y-%m-%d") if date_depart else None,
                # Profil de performance individuel
                "profile": {
                    "base_auto": random.uniform(-1.2, 1.2),
                    "base_rig": random.uniform(-1.2, 1.2),
                    "base_com": random.uniform(-1.2, 1.2),
                    "base_eng": random.uniform(-1.2, 1.2),
                    "variance": random.uniform(0.3, 0.8),
                    "trend": random.uniform(-0.15, 0.15),
                },
            }
        )

    return students


def generate_seances(n: int, start: date, creneau: tuple[str, str]) -> list[dict]:
    """Génère n séances hebdomadaires à partir de start."""
    heure_debut, heure_fin = creneau
    seances = []
    d = start
    for i in range(1, n + 1):
        seances.append({"num": i, "date": d, "heure_debut": heure_debut, "heure_fin": heure_fin})
        d += timedelta(weeks=1)
    return seances


def build_releves(students: list[dict], seance: dict) -> list[dict]:
    """Construit les relevés d'une séance pour les étudiants encore actifs."""
    releves = []
    for s in students:
        if s["date_depart"] and s["date_depart"] < seance["date"].strftime("%Y-%m-%d"):
            continue

        p = s["profile"]
        idx = seance["num"] - 1
        avg_base = (p["base_auto"] + p["base_rig"] + p["base_com"] + p["base_eng"]) / 4

        commentaire: str | None = None
        if random.random() < 0.25:
            pool = COMMENTAIRES_POS if avg_base >= 0.5 else (
                COMMENTAIRES_NEG if avg_base <= -0.5 else COMMENTAIRES_POS + COMMENTAIRES_NEG
            )
            commentaire = random.choice(pool)

        releves.append(
            {
                "nom": s["nom"],
                "presence": gen_presence(seance["heure_debut"]),
                "autonomie": gen_score(p["base_auto"], p["variance"], p["trend"], idx),
                "rigueur": gen_score(p["base_rig"], p["variance"], p["trend"], idx),
                "communication": gen_score(p["base_com"], p["variance"], p["trend"], idx),
                "engagement": gen_score(p["base_eng"], p["variance"], p["trend"], idx),
                "commentaire": commentaire,
            }
        )
    return releves


# ---------------------------------------------------------------------------
# Écriture XLSX
# ---------------------------------------------------------------------------


def write_config_sheet(ws, projet: str, groupe: str, students: list[dict]) -> None:
    """Écrit la feuille Config (métadonnées projet + liste étudiants)."""
    ws["A1"], ws["B1"] = "Projet", projet
    ws["A2"], ws["B2"] = "Groupe", groupe

    headers = ["Nom", "INE", "Anonyme", "Pseudo", "Date de départ"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = h
        cell.font = Font(bold=True)

    for i, s in enumerate(students):
        row = 5 + i
        ws.cell(row=row, column=1).value = s["nom"]
        ws.cell(row=row, column=2).value = s["ine"]
        ws.cell(row=row, column=3).value = s["anonyme"]
        ws.cell(row=row, column=4).value = s["pseudo"]
        ws.cell(row=row, column=5).value = s["date_depart"]


def write_seance_sheet(
    ws,
    seance_num: int,
    date_seance: date,
    heure_debut: str,
    heure_fin: str,
    enseignant: str,
    releves: list[dict],
) -> None:
    """Écrit une feuille de séance au format attendu par l'importeur."""
    # Ligne 1 : titre fusionné
    ws.merge_cells("A1:J1")
    ws["A1"] = "COMPAS — Relevé de séance"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Ligne 2 : métadonnées
    ws["A2"] = "Séance n°"
    ws["B2"] = seance_num
    ws["C2"] = "Date"
    ws["D2"] = date_seance.strftime("%d/%m/%Y")
    ws["E2"] = "Heure début"
    ws["F2"] = heure_debut
    ws["G2"] = "Enseignant"
    ws["H2"] = enseignant
    ws["I2"] = "Heure fin"
    ws["J2"] = heure_fin

    # Ligne 3 : vide (espacement)

    # Ligne 4 : en-têtes colonnes
    col_headers = [
        "Étudiant",
        "Présence",
        "Autonomie",
        "Rigueur",
        "Communication",
        "Engagement",
        "Commentaire",
    ]
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = h
        cell.font = Font(bold=True)

    # Ligne 5 : rappel symboles
    ws.cell(row=5, column=1).value = "(symboles)"
    ws.cell(row=5, column=2).value = "P / A / R:N / D:Xh / N"
    ws.cell(row=5, column=3).value = "-2 … +2"

    # Lignes 6+ : données
    for i, r in enumerate(releves):
        row = 6 + i
        ws.cell(row=row, column=1).value = r["nom"]
        ws.cell(row=row, column=2).value = r["presence"]
        ws.cell(row=row, column=3).value = r["autonomie"]
        ws.cell(row=row, column=4).value = r["rigueur"]
        ws.cell(row=row, column=5).value = r["communication"]
        ws.cell(row=row, column=6).value = r["engagement"]
        ws.cell(row=row, column=7).value = r["commentaire"]


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------


def generate(
    out_path: Path,
    csv_path: Path | None,
    projet: str,
    groupe: str,
    enseignant: str,
    students: list[dict],
    seances: list[dict],
) -> None:
    """Génère le fichier XLSX et, si demandé, le CSV récapitulatif."""
    wb = openpyxl.Workbook()

    ws_config = wb.active
    ws_config.title = "Config"
    write_config_sheet(ws_config, projet, groupe, students)

    # Feuille Modèle vide (ignorée par l'import, présente dans le template Excel)
    wb.create_sheet("Modèle")

    csv_rows: list[dict] = []

    for seance in seances:
        releves = build_releves(students, seance)
        ws = wb.create_sheet(f"S{seance['num']}")
        write_seance_sheet(
            ws,
            seance_num=seance["num"],
            date_seance=seance["date"],
            heure_debut=seance["heure_debut"],
            heure_fin=seance["heure_fin"],
            enseignant=enseignant,
            releves=releves,
        )

        if csv_path:
            for r in releves:
                student_meta = next(s for s in students if s["nom"] == r["nom"])
                csv_rows.append(
                    {
                        "projet": projet,
                        "groupe": groupe,
                        "enseignant": enseignant,
                        "seance": seance["num"],
                        "date": seance["date"].strftime("%d/%m/%Y"),
                        "heure_debut": seance["heure_debut"],
                        "heure_fin": seance["heure_fin"],
                        "nom": r["nom"],
                        "ine": student_meta["ine"],
                        "anonyme": student_meta["anonyme"],
                        "presence": r["presence"],
                        "autonomie": r["autonomie"],
                        "rigueur": r["rigueur"],
                        "communication": r["communication"],
                        "engagement": r["engagement"],
                        "commentaire": r["commentaire"] or "",
                    }
                )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"XLSX  : {out_path}  ({len(students)} étudiants, {len(seances)} séances)")

    if csv_path and csv_rows:
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=csv_rows[0].keys())
            writer.writeheader()
            writer.writerows(csv_rows)
        print(f"CSV   : {csv_path}  ({len(csv_rows)} lignes)")


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Génère des données de test réalistes pour COMPAS.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--out",
        default="data/test_generated.xlsx",
        help="Chemin du fichier XLSX à produire",
    )
    parser.add_argument(
        "--csv",
        default=None,
        metavar="PATH",
        help="Chemin du fichier CSV récapitulatif (optionnel)",
    )
    parser.add_argument(
        "--fixture",
        action="store_true",
        help="Met à jour tests/fixtures/test_projet.xlsx (seed=0, 6 étudiants, 3 séances)",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=None,
        help="Graine aléatoire pour des données reproductibles",
    )
    parser.add_argument(
        "--students",
        type=int,
        default=15,
        help="Nombre d'étudiants à générer",
    )
    parser.add_argument(
        "--sessions",
        type=int,
        default=6,
        help="Nombre de séances à générer",
    )
    args = parser.parse_args()

    # Mode fixture : paramètres fixes, chemin imposé
    if args.fixture:
        args.seed = args.seed if args.seed is not None else 0
        args.students = 6
        args.sessions = 3
        args.out = "tests/fixtures/test_generated.xlsx"

    if args.seed is not None:
        random.seed(args.seed)
        Faker.seed(args.seed)

    fake = Faker("fr_FR")

    projet = random.choice(PROJETS)
    groupe = random.choice(GROUPES)
    enseignant = random.choice(ENSEIGNANTS)
    creneau = random.choice(CRENEAUX)

    students = generate_students(fake, args.students)

    # Premier jour : ~ (sessions + 2) semaines avant aujourd'hui, aligné sur un lundi
    start = date.today() - timedelta(weeks=args.sessions + 2)
    start -= timedelta(days=start.weekday())

    seances = generate_seances(args.sessions, start, creneau)

    print(f"Projet    : {projet} ({groupe})")
    print(f"Enseignant: {enseignant}  |  Créneau : {creneau[0]}–{creneau[1]}")

    out_path = Path(args.out)
    csv_path = Path(args.csv) if args.csv else None

    generate(out_path, csv_path, projet, groupe, enseignant, students, seances)


if __name__ == "__main__":
    main()
